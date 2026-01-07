#   Copyright (C) 2025-2026 Credit Mutuel Arkea
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.
#
"""
Export a LangFuse dataset experiment (run) results.

Usage:
    export_experiments.py [-v] --json-config-file=<jcf>

Description:
    This script is used to export one or more dataset experiments specified on a json configuration file.
    The configuration file specifies the export template:
        - For "xlsx", see examples/dataset-items-example.xlsx
        - For "json", see examples/dataset-items-example.json

Arguments:
    --json-config-file=<jcf>   Path to the input config file. This is a required argument.

Options:
    -v                         Enable verbose output for debugging purposes. If not set, the script runs silently except for errors.
    -h, --help                 Display this help message and exit.
    --version                  Display the version of the script.

Examples:
    python export_experiments.py --json-config-file=path/to/config-file.json
"""

import os
import re
from datetime import datetime

from docopt import docopt
from langfuse import Langfuse
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from scripts.common.logging_config import configure_logging
from scripts.common.models import ActivityStatus, StatusWithReason
from scripts.dataset.export.model import (
    DatasetExperimentItem,
    DatasetExperimentItemRun,
    DatasetExperimentScore,
    ExportExperimentsInput,
    ExportExperimentsOutput,
)

from gen_ai_orchestrator.services.security.security_service import (
    fetch_secret_key_value,
)


# LangFuse-specific functions
def fetch_trace_by_item_and_dataset_run(client, dataset_run, item):
    """
    Fetches the trace for a dataset item from a LangFuse dataset run.

    Args:
        dataset_run: The dataset run with items.
        item: The dataset item.

    Returns:
        Trace data if found, otherwise None.
    """
    for item_run in dataset_run:
        if item.id == item_run.dataset_item_id:
            trace = client.api.trace.get(item_run.trace_id)
            return trace
    return None


def append_runs_langfuse(
    client, dataset_experiments, dataset_item
) -> DatasetExperimentItem:
    runs: list[DatasetExperimentItemRun] = []

    for run_name in dataset_experiments.experiment_names:
        dataset_run = client.api.datasets.get_run(
            dataset_name=dataset_experiments.dataset_name,
            run_name=run_name,
        )
        trace = fetch_trace_by_item_and_dataset_run(
            client, dataset_run.dataset_run_items, dataset_item
        )
        runs.append(
            DatasetExperimentItemRun(
                output=trace.output,
                scores=list(
                    map(
                        lambda s: DatasetExperimentScore(
                            name=s.name,
                            value=s.value,
                            comment=s.comment,
                            trace_id=s.trace_id,
                        ),
                        trace.scores,
                    )
                ),
                trace_id=trace.id,
                metadata=dataset_run.metadata,
            )
        )

    return DatasetExperimentItem(
        input=dataset_item.input,
        expected_output=dataset_item.expected_output,
        metadata=dataset_item.metadata,
        runs=runs,
    )


def create_excel_output(
    iterations: list[str],
    items: list[DatasetExperimentItem],
    output_file: str,
    metric_names: list[str],
):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    wb = load_workbook(
        os.path.join(
            script_dir, '../../../examples/generate_dataset_input.example.xlsx'
        )
    )
    sheet = wb['Template_Suivi_Recette']

    # Prepare NO_RAG stat if asked
    no_rag_percentages = {}
    if 'NoRagStat' in metric_names:
        for run_idx, run_name in enumerate(iterations):
            no_rag_count = 0
            for item in items:
                answer = item.runs[run_idx].output.get('answer')
                if answer == 'NO_RAG_ANSWER':
                    no_rag_count += 1
            if items:
                no_rag_percentages[
                    run_idx
                ] = f"{round(no_rag_count / len(items) * 100)}%"
            else:
                no_rag_percentages[run_idx] = '0%'

    for i in range(len(iterations)):
        start_row = 7 + 6 * i
        sheet.merge_cells(
            start_row=start_row, start_column=2, end_row=12 + 6 * i, end_column=2
        )
        sheet[f"B{start_row}"] = iterations[i]
        sheet[f"C{start_row}"] = items[0].runs[i].metadata['llm']['model']
        sheet[f"D{start_row}"] = items[0].runs[i].metadata['llm']['temperature']
        sheet[f"E{start_row}"] = (
            no_rag_percentages.get(i, 'N/A') if 'NoRagStat' in metric_names else ''
        )
        sheet[f"F{start_row}"] = items[0].runs[0].metadata['k']
        sheet[f"G{start_row}"] = items[0].runs[i].metadata['document_index_name']

    for i in range(len(items)):
        col_letter = get_column_letter(
            10 + i
        )  # Start at col J (index 9 as indexes starts at 0 for letter A)
        sheet[f"{col_letter}3"] = items[i].metadata.get('topic', '')
        sheet[f"{col_letter}4"] = items[i].input.get('question', '')
        sheet[f"{col_letter}5"] = items[i].expected_output.get('answer', '')
        for j in range(len(iterations)):
            start_row = 7 + 6 * j
            sheet[f"{col_letter}{start_row}"] = (
                items[i].runs[j].output.get('answer', '')
            )
            sheet[f"{col_letter}{start_row + 1}"] = '\n\n'.join(
                [
                    format_document_for_excel(doc)
                    for doc in items[i].runs[j].output.get('documents', [])
                    if isinstance(doc, dict)
                ]
            )
            sheet[f"{col_letter}{start_row + 5}"] = '\n\n'.join(
                f"{s.name} : {s.value:.2f} ({s.comment.split(':', 1)[1].strip()})"
                if ':' in s.comment
                else f"{s.name} : {s.value:.2f}"
                for s in items[i].runs[j].scores
                if s.name in metric_names and s.name != 'NoRagStat'
            )

    wb.save(output_file)


ILLEGAL_EXCEL_CHARS = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')


def sanitize_for_excel(text: str) -> str:
    # 1. Delete ```markdown
    if text.startswith('```'):
        text = text.strip('`').replace('markdown\n', '', 1).strip()

    # 2. Delete illegal excel chars
    text = ILLEGAL_EXCEL_CHARS.sub('', text)

    return text


def format_document_for_excel(doc: dict) -> str:
    """
    Format a document as:
    [id: title]
    sanitized_page_content
    """
    metadata = doc.get('metadata', {})

    doc_id = metadata.get('id', 'unknown-id')
    title = metadata.get('title', 'untitled')

    raw_content = doc.get('page_content', '')
    clean_content = sanitize_for_excel(raw_content)

    return f"[{doc_id}: {title}]\n{clean_content}"


def timestamped_filename(filename: str) -> str:
    base, ext = os.path.splitext(filename)
    timestamped = datetime.now().strftime('%Y%m%d-%H%M%S')
    return f"{base}-{timestamped}{ext}"


def main():
    start_time = datetime.now()
    cli_args = docopt(__doc__, version='Run Experiment 1.0.0')
    logger = configure_logging(cli_args)

    dataset_name: str = ''
    items: list[DatasetExperimentItem] = []
    try:
        logger.info('Loading input data...')
        input_config = ExportExperimentsInput.from_json_file(
            cli_args['--json-config-file']
        )
        logger.debug(f"\n{input_config.format()}")

        location = f"{input_config.bot.file_location}/{input_config.bot.namespace}-{input_config.bot.bot_id}/output"
        template_file_path = (
            f"{location}/{timestamped_filename(input_config.template.file)}"
        )

        client = Langfuse(
            host=str(input_config.observability_setting.url),
            public_key=input_config.observability_setting.public_key,
            secret_key=fetch_secret_key_value(
                input_config.observability_setting.secret_key
            ),
        )
        dataset_name = input_config.dataset_experiments.dataset_name
        dataset = client.get_dataset(name=dataset_name)
        items: list[DatasetExperimentItem] = []
        for item in dataset.items:
            items.append(
                append_runs_langfuse(client, input_config.dataset_experiments, item)
            )

        if 'xlsx' == input_config.template.type:
            create_excel_output(
                iterations=input_config.dataset_experiments.experiment_names,
                items=items,
                output_file=template_file_path,
                metric_names=input_config.metric_names,
            )
        else:
            raise RuntimeError(
                f"The '{input_config.dataset.template.type}' dataset template is not yet supported!"
            )

        activity_status = StatusWithReason(status=ActivityStatus.COMPLETED)
    except Exception as e:
        full_exception_name = f"{type(e).__module__}.{type(e).__name__}"
        activity_status = StatusWithReason(
            status=ActivityStatus.FAILED, status_reason=f"{full_exception_name} : {e}"
        )
        logger.error(e, exc_info=True)

    output = ExportExperimentsOutput(
        status=activity_status,
        dataset_name=dataset_name,
        duration=datetime.now() - start_time,
        items_count=len(items),
        success_rate=100,
    )
    logger.debug(f"\n{output.format()}")


if __name__ == '__main__':
    main()
